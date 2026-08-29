import pandas as pd
import matplotlib.pyplot as plt
import openpyxl

from openpyxl.chart import Reference, PieChart

def askUser():
    total = 0

    # This loop runs five times, asks the user for a number, and adds each number to the total.
    for i in range(5):
        number = float(input("Enter a number: "))
        total += number

    print("The total is:", total)

def askIncome():
    # This loop runs five times and asks the user for a name and annual income.
    # Each person's name and income is appended to the existing final.csv file.
    for i in range(5):
        name = input("Enter the person's name: ")
        income = input("Enter the person's annual income: ")

        # Open final.csv in append mode so the existing four entries are not erased.
        with open(r"C:\FinalExam\final.csv", "a") as file:
            file.write("\n" + name + "," + income)

# Create an Excel file containing the data and an Excel pie chart.
def excelPie():
    # Read the existing final.csv file into a pandas DataFrame.
    # The CSV file does not have headers, so we provide the column names.
    df = pd.read_csv(r"C:\FinalExam\final.csv", header=None, names=["Name", "Income"])

    # Convert the income values to integers so Excel can use them in the pie chart.
    df["Income"] = df["Income"].astype(int)

    # Create a new Excel workbook using the data from final.csv.
    df.to_excel("final.xlsx", index=False)

    # Open the newly created Excel workbook.
    wb = openpyxl.load_workbook("final.xlsx")

    # Select the active worksheet in the workbook.
    ws = wb.active

    # Create a PieChart object for the Excel worksheet.
    myChart = PieChart()

    # Select the income data from column B, including the header.
    data = Reference(ws, min_col=2, min_row=1, max_row=ws.max_row)

    # Select the names from column A to use as the pie chart labels.
    labels = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)

    # Add the income data to the pie chart.
    myChart.add_data(data, titles_from_data=True)

    # Set the names as the labels for the pie chart slices.
    myChart.set_categories(labels)

    # Set the title of the pie chart to the StudentID and today's date.
    myChart.title = "AMAHOP2918 August 28, 2026"

    # Add the pie chart to the worksheet starting at cell D2.
    ws.add_chart(myChart, "D2")

    # Save the Excel workbook containing the data and pie chart.
    wb.save("final.xlsx")


# Create a vertical bar graph using matplotlib.
def verticalBar():
    # Read the data from final.csv into a pandas DataFrame.
    # The CSV file does not have headers, so we provide the column names.
    df = pd.read_csv("final.csv", names=["Name", "Income"])

    # Convert the income values to integers for the bar graph.
    df["Income"] = df["Income"].astype(int)

    # Create a vertical bar graph using names on the x-axis and income on the y-axis.
    df.plot(x="Name", y="Income", kind="bar")

    # Set the title of the bar graph to the StudentID and today's date.
    plt.title("AMAHOP2918 August 28, 2026")

    # Label the horizontal axis.
    plt.xlabel("Name")

    # Label the vertical axis.
    plt.ylabel("Annual Income")

    # Rotate the names so they are easier to read.
    plt.xticks(rotation=45)

    # Adjust the spacing so all labels fit on the graph.
    plt.tight_layout()

    # Display the vertical bar graph.
    plt.show()

def main():
    askUser()
    askIncome()
    excelPie()
    verticalBar()

if __name__ == "__main__":
    main()
